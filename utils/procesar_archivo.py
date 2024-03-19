import openpyxl
import pandas as pd
import jaro
from utils import consume
from os import getcwd, path
import os
import pickle
import shutil


url ='./files/'
url2 ="./files2/"
url3 = './files/dummy.pkl'
url4 = "./files2/dummy.pkl"
ext = ".pkl"
url_deparments = "./departments/"

# para buscar listas 
def comprobar2(name:str):
    
    book= openpyxl.load_workbook(url2+name, data_only=True)
    hoja = book.active
    celdas_id = hoja['A2':'A100']
    celdas_nombre = hoja['B2':'B100']
    lista_cargue = []
    identificacion = []
    nombre = []

    for fila in celdas_id:
        empleado = [celda.value for celda in fila]
        empleado_id = str(empleado[0])
        empleado_id = empleado_id.upper()   
        if empleado_id !='NONE':
            identificacion.append(empleado_id)
    for fila in celdas_nombre:
        empleado = [celda.value for celda in fila]
        empleado_nombre = str(empleado[0])
        empleado_nombre = empleado_nombre.upper()
        if empleado_nombre !='NONE':
            nombre.append(empleado_nombre)
    for i in range(len(nombre)):
        lista_cargue.append((identificacion[i], nombre[i]))
    
    
    dcargue= pd.DataFrame(lista_cargue, columns=['Id','Nombre'])
    #almacenar datos en la base de datos sql
    dcargue.to_pickle(url4)
    datos = pd.read_pickle(url4) 
    lista = datos.to_numpy().tolist()
    lista1 = consume.consumir_2(lista,name)
    return lista1

def comprobar3(name:str,lists:list,coincidencia:int):
    
    book= openpyxl.load_workbook(url2+name, data_only=True)
    hoja = book.active
    celdas_id = hoja['A2':'A500']
    celdas_nombre = hoja['B2':'B500']
    lista_cargue = []
    identificacion = []
    nombre = []

    for fila in celdas_id:
        empleado = [celda.value for celda in fila]
        empleado_id = str(empleado[0])
        empleado_id = empleado_id.upper()   
        if empleado_id !='NONE':
            identificacion.append(empleado_id)
    for fila in celdas_nombre:
        empleado = [celda.value for celda in fila]
        empleado_nombre = str(empleado[0])
        empleado_nombre = empleado_nombre.upper()
        if empleado_nombre !='NONE':
            nombre.append(empleado_nombre)
    for i in range(len(nombre)):
        lista_cargue.append((identificacion[i], nombre[i]))
    
    
    dcargue= pd.DataFrame(lista_cargue, columns=['Id','Nombre'])
    #almacenar datos en la base de datos sql
    dcargue.to_pickle(url4)
    datos = pd.read_pickle(url4) 
    lista = datos.to_numpy().tolist()
    lista1 = consume.consumir_2(lista,name,coincidencia,lists)
    return lista1
    


def buscar2_id(name:str,coincidencia:int):
    coincidencia = coincidencia/100
    try:
        files = open(url4)
        files.close()    
        datos = pd.read_pickle(url4) 
        lista = datos.to_numpy().tolist()
        name = name.upper()
        val=''
        for datos in lista :
            datos = str(datos[0])
            p= jaro.jaro_metric(name,datos)
            if p >= coincidencia:
                val='x(identificacion)'
    
    except FileNotFoundError:
        val=''  
    return val

def buscar2_nombre(name:str,coincidencia:int):
    coincidencia = coincidencia/100
    try:
        files = open(url4)
        files.close()    
        datos = pd.read_pickle(url4) 
        lista = datos.to_numpy().tolist()
        name = name.upper()
        val=''
        for datos in lista :
            datos = datos[1]
            p= jaro.jaro_metric(name,datos)
            if p >= coincidencia :
                val='x(Nombre)'
    
    except FileNotFoundError:
        val=''
    return val

# #Recibe el nombre del archivo para archivarlo en un pickle
# def comprobar(name:str):
#     book= openpyxl.load_workbook(url+name, data_only=True)
#     hoja = book.active
#     celdas = hoja['A2':'A1000']
#     lista_cargue = []
#     for fila in celdas:
#         empleado = [celda.value for celda in fila]
#         empleado= str(empleado[0])
#         empleado = empleado.upper()
#         lista_cargue.append(empleado)
#     dcargue= pd.DataFrame(lista_cargue, columns=['first_name'])
#     #almacenar datos en la base de datos sql
#     dcargue.to_pickle(url3)

#Recibe el nombre del archivo para archivarlo en un pickle
def comprobar(name:str):
    book= openpyxl.load_workbook(url+name, data_only=True)
    hoja = book.active
    celdas = hoja['A2':'A1001']
    celdas2 = hoja['B2':'B1001']
    lista_cargue = []
    filename =os.path.splitext(name)[0] + '.pkl'
    for fila,fila2 in zip(celdas,celdas2):
        identificacion = [celda.value for celda in fila]
        nombre = [celda.value for celda in fila2]
        if identificacion[0]!=None and nombre[0]!=None:
            identificacion= str(identificacion[0])
            nombre= str(nombre[0])
            nombre = nombre.upper()
            lista_cargue.append((identificacion,nombre))
    dcargue= pd.DataFrame(lista_cargue, columns=['Identificacion','Firstname'])
    #almacenar datos en la base de datos sql
    dcargue.to_pickle(url+filename)
    return filename


#Busca Persona en listas
def buscar(name:str,coincidencia:int, lists:list):
    coincidencia = coincidencia/100
    lista_tables = []
    listas_encontrado=[]
    for item in lists:
       lista_tables.append(item.descripcion)

    for name2 in lista_tables:
        existe= path.exists(url+name2)
        if existe:
            datos = pd.read_pickle(url+name2) 
            lista = datos.to_numpy().tolist()
            name = name.upper()
            indice = 2
            for datos in lista :
                datos=str(datos[0])
                p= jaro.jaro_metric(name,datos)
                if p>=coincidencia : 
                    name2 = name2.replace(".pkl", "")
                    listas_encontrado.append(name2+' por Identificacion('+str(indice)+')')
                indice+=1
            indice2 = 2 
            for datos in lista :
                datos=str(datos[1])
                p= jaro.jaro_metric(name,datos)
                if p>=coincidencia : 
                    name2 = name2.replace(".pkl", "")
                    listas_encontrado.append(name2+' por Nombre('+str(indice2)+') ')
                indice2+=1     
        else:
            name2 = name2.replace(".pkl", "")
            listas_encontrado.append(name2+' No existe el archivo')
    return listas_encontrado

#Busca Persona en listas
def buscar2(name:str,coincidencia:int, lists:list):
    coincidencia = coincidencia/100
    lista_tables = []
    listas_encontrado=[]
    for item in lists:
        lista_tables.append(item.descripcion)

    for name2 in lista_tables:
        existe= path.exists(url+name2)
        if existe:
            datos = pd.read_pickle(url+name2) 
            lista = datos.to_numpy().tolist()
            name = name.upper()
            for datos in lista :
                datos=str(datos[0])
                p= jaro.jaro_metric(name,datos)
                if p>=coincidencia : 
                    name2 = name2.replace(".pkl", "")
                    listas_encontrado.append(name2+' por Identificacion')
            for datos in lista :
                datos=str(datos[1])
                p= jaro.jaro_metric(name,datos)
                if p>=coincidencia : 
                    name2 = name2.replace(".pkl", "")
                    listas_encontrado.append(name2+' por Nombre')
        else:
            # name2 = name2.replace(".pkl", "")
            # listas_encontrado.append(name2+' No existe el archivo')
            pass 
    return listas_encontrado

#busca persona recibiendo la coincidencia
def buscar_listas_person(name:str,coincidencia:int):
    datos = pd.read_pickle(url3) 
    lista1=[]
    lista = datos.to_numpy().tolist()
    name = name.upper()
    val=False
    for datos in lista :
        datos=str(datos)
        p= jaro.jaro_metric(name,datos)
        if p>=0.90 :
            val=True
    lista1.append(val)
    return lista1

#busca el pkl y retorna lista de palabras claves
def words_keys(name:str):
    archivo=url_deparments+name+".pkl"
    datos = pd.read_pickle(archivo)
    datos = datos.to_numpy().tolist()
    if len(datos) == 0:
        datos.append({'departmen': 'default', 'words': 'default'})
    else:
        datos = {'departmen':name,'words':datos[0]}
    
    return datos

#Crea un departamento
def create_department(name):
    archivo=url_deparments+name+".pkl"
    mi_lista={}
    datos = pd.DataFrame(mi_lista)
    datos.to_pickle(archivo)
    return True
    

#Añade palabra clave a departamento
def add_word_key(archivo:str,word:str):
    archivo1 = archivo
    archivo=url_deparments+archivo+".pkl"
    lista = pd.read_pickle(archivo)
    lista = lista.to_numpy().tolist()
    if len(lista) == 0:
        lista.append(word)
    else:
        lista[0].append(word)
    os.remove(archivo)
    datos = pd.DataFrame(lista)
    datos.to_pickle(archivo)
    return True

def get_departments(): 
    archivos_pkl_lista = [archivo[:-4] for archivo in os.listdir(url_deparments) if archivo.endswith('.pkl')]
    return archivos_pkl_lista